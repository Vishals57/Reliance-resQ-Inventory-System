import pandas as pd
import os
import uuid
import zipfile
from datetime import datetime

DB_FILE = "resQ_Enterprise_Inventory.xlsx"
BACKUP_DIR = "Backups"
# Backward compatible default list (used only to seed the Engineers sheet once).
ENGINEERS = ["Yogesh Bhosale", "Sidram Battul", "Sham Kachi", "sameer shaikh", "Rajesh chavan", "Sahil Dighe", "Gajanan gawande", "Kiran misal", "kishor panchal", "Raju Yadav", "Lalemashak kunnure", "Faruk Shaikh", "Ashraf Momin", "Abhitabh Gupta", "Prahalad Yadav", "Abdul Rehman", "Ishwar Panchal", "Nagendra Yadav", "Pramod Valekar", "Rahim Dindore"]

ENGINEER_SHEET = "Engineers"
ENGINEER_COLUMNS = ["Engineer_Name", "BP_ID", "PPRR_ID", "Active", "Created_At"]

# Post–OUTWARD billing / TCR workflow (matches your tracking sheet columns)
SERVICE_JOB_SHEET = "Service_Billing"
SERVICE_JOB_COLUMNS = [
    "Row_ID",
    "Sr_No",
    "Engineer_Name",
    "Date",
    "Job_Card_Invoice_No",
    "Outward_Date",
    "TCR_No",
    "TCR_Date",
    "Warranty_Warranty_No",
    "Bill_Amount",
    "Incentive",
    "Ticket_Closed_On_Date",
    "Part_Name_Part_No",
    "Material_Description",
    "HSN_Code",
    "Quantity",
    "MRP_Product_Price",
    "CGST",
    "SGST",
    "Status",
    "Remarks",
    "Article_No",
    "Part_Sr_No",
    "Money_Received",
    "Created_At",
    "Updated_At",
]

MASTER_COLUMNS = ["Article_No", "Part_Name", "CP", "SP", "Category", "Stock_Level"]
# `Date` is kept as a generic "last update" timestamp for backward compatibility,
# while `In_Date` and `Out_Date` store movement timestamps explicitly.
TRANSACTION_COLUMNS = [
    "Article_No",
    "Sr_No",
    "Purchase_Type",
    "Tax_Invoice_No",
    "Status",
    "Engineer",
    "Charges",
    "In_Date",
    "Out_Date",
    "Date",
]


def _read_sheet_df(sheet_name: str, columns: list):
    try:
        df = pd.read_excel(DB_FILE, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = None
    return df[columns]


def _ensure_job_schema(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    df = df.copy()
    for col in SERVICE_JOB_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[SERVICE_JOB_COLUMNS]


def save_all_workbook(df_master: pd.DataFrame, df_trans: pd.DataFrame, df_eng=None, df_jobs=None):
    """Write Master, Transactions, Engineers, Service_Billing together (never drop sheets)."""
    if df_eng is None:
        df_eng = _read_sheet_df(ENGINEER_SHEET, ENGINEER_COLUMNS)
    else:
        df_eng = _ensure_eng_schema(df_eng)
    if df_jobs is None:
        df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    else:
        df_jobs = _ensure_job_schema(df_jobs)

    df_trans = _ensure_schema(df_trans)

    with pd.ExcelWriter(DB_FILE, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master", index=False)
        df_trans.to_excel(writer, sheet_name="Transactions", index=False)
        df_eng.to_excel(writer, sheet_name=ENGINEER_SHEET, index=False)
        df_jobs.to_excel(writer, sheet_name=SERVICE_JOB_SHEET, index=False)

    apply_service_billing_green_rows()
    apply_excel_formatting()


def _ensure_backup_dir() -> str:
    """Ensure the backup directory exists and return its path."""
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR, exist_ok=True)
    return BACKUP_DIR


def backup_database(include_qr_folder: bool = True):
    """Create a timestamped ZIP backup of the database and optional QR assets."""
    migrate_db()
    if not os.path.exists(DB_FILE):
        return False, "Database file not found."

    backup_root = _ensure_backup_dir()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"resQ_backup_{stamp}.zip"
    backup_path = os.path.join(backup_root, backup_name)

    try:
        with zipfile.ZipFile(backup_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(DB_FILE, os.path.basename(DB_FILE))
            if include_qr_folder and os.path.isdir("Box_QRs"):
                for root, _, files in os.walk("Box_QRs"):
                    for file in files:
                        src_path = os.path.join(root, file)
                        rel_path = os.path.relpath(src_path, os.path.dirname(DB_FILE) or ".")
                        archive.write(src_path, rel_path)

        return True, backup_path
    except Exception as exc:
        return False, f"Backup failed: {exc}"


def _ensure_eng_schema(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=ENGINEER_COLUMNS)
    df = df.copy()
    for col in ENGINEER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[ENGINEER_COLUMNS]


def apply_service_billing_green_rows():
    """
    After Excel save: header row yellow; data row light green when TCR_No is filled AND Money_Received is true.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except Exception:
        return

    if not os.path.exists(DB_FILE):
        return
    try:
        wb = load_workbook(DB_FILE)
    except Exception:
        return
    if SERVICE_JOB_SHEET not in wb.sheetnames:
        return

    ws = wb[SERVICE_JOB_SHEET]
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for cell in ws[1]:
        cell.fill = yellow

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    def col_idx(name: str):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    ci_tcr = col_idx("TCR_No")
    ci_money = col_idx("Money_Received")
    if not ci_tcr or not ci_money:
        wb.save(DB_FILE)
        return

    def money_ok(val) -> bool:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return False
        s = str(val).strip().lower()
        if s in ("", "nan", "none"):
            return False
        return s in ("true", "1", "yes", "y", "received", "paid")

    for r in range(2, ws.max_row + 1):
        tcr = ws.cell(row=r, column=ci_tcr).value
        money = ws.cell(row=r, column=ci_money).value
        has_tcr = bool(str(tcr).strip()) if tcr is not None and str(tcr).strip().lower() != "nan" else False
        row_ok = has_tcr and money_ok(money)
        fill = green if row_ok else white
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

    try:
        wb.save(DB_FILE)
    except Exception:
        pass


def apply_excel_formatting():
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
    except Exception:
        return

    if not os.path.exists(DB_FILE):
        return

    try:
        wb = load_workbook(DB_FILE)
    except Exception:
        return

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.row_dimensions[1].height = 26
            for cell in ws[1]:
                if cell.value is None:
                    continue
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            
            # Start with header width
            header_cell = ws[f'{column_letter}1']
            if header_cell.value is not None:
                max_length = len(str(header_cell.value)) + 2  # Add padding for header
            
            # Check all content in column
            for cell in column_cells[1:]:  # Skip header
                if cell.value is None:
                    continue
                value_str = str(cell.value).strip()
                value_length = len(value_str)
                if value_length > max_length:
                    max_length = value_length
            
            # Calculate final width with better padding and min/max
            # Min 12, Max 50 to ensure all text is visible
            adjusted_width = min(max(12, max_length + 3), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(DB_FILE)
    except Exception:
        pass


def lookup_engineer_bp_pprr(engineer_name: str):
    df = _read_sheet_df(ENGINEER_SHEET, ENGINEER_COLUMNS)
    if df.empty or not engineer_name:
        return "", ""
    key = str(engineer_name).strip().lower()
    df = df.copy()
    df["_k"] = df["Engineer_Name"].astype(str).str.strip().str.lower()
    m = df[df["_k"] == key]
    if m.empty:
        return "", ""
    row = m.iloc[0]
    bp = "" if row.get("BP_ID") is None else str(row.get("BP_ID")).strip()
    pprr = "" if row.get("PPRR_ID") is None else str(row.get("PPRR_ID")).strip()
    return bp, pprr


def append_service_job_after_outward(df_jobs: pd.DataFrame, article_no: str, part_sr_no: int, engineer_name: str, part_name: str, outward_dt: str):
    df_jobs = _ensure_job_schema(df_jobs)
    serial = 1
    if not df_jobs.empty:
        try:
            serial = int(pd.to_numeric(df_jobs["Sr_No"], errors="coerce").fillna(0).max()) + 1
        except Exception:
            serial = len(df_jobs) + 1

    # Get invoice number and additional data from transactions
    invoice_no = ""
    material_desc = ""
    try:
        df_trans = _read_sheet_df("Transactions", TRANSACTION_COLUMNS)
        if not df_trans.empty:
            # Find the most recent transaction for this article
            article_matches = df_trans[df_trans["Article_No"].astype(str).str.strip() == str(article_no).strip()]
            if not article_matches.empty:
                # Get the most recent transaction (by Sr_No or Date)
                latest_trans = article_matches.sort_values("Sr_No", ascending=False).iloc[0]
                invoice_no = str(latest_trans.get("Tax_Invoice_No", "")).strip()
                # Auto-generate material description from purchase type
                purchase_type = str(latest_trans.get("Purchase_Type", "")).strip()
                if purchase_type:
                    material_desc = f"Service work on {purchase_type.lower()} purchase"
    except Exception:
        pass  # If lookup fails, keep empty

    # Get part details from master catalog
    try:
        df_master = _read_sheet_df("Master", MASTER_COLUMNS)
        if not df_master.empty:
            master_matches = df_master[df_master["Article_No"].astype(str).str.strip() == str(article_no).strip()]
            if not master_matches.empty:
                master_row = master_matches.iloc[0]
                if not part_name:  # If part_name not provided, get from master
                    part_name = str(master_row.get("Part_Name", "")).strip()
                # Enhance material description with part info
                if material_desc and part_name:
                    material_desc += f" - {part_name}"
                elif not material_desc and part_name:
                    material_desc = f"Service work on {part_name}"
    except Exception:
        pass

    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rid = str(uuid.uuid4())

    row = {
        "Row_ID": rid,
        "Sr_No": serial,
        "Engineer_Name": engineer_name,
        "Date": today,
        "Job_Card_Invoice_No": invoice_no,  # Now populated from transaction
        "Outward_Date": outward_dt,
        "TCR_No": "",
        "TCR_Date": "",
        "Warranty_Warranty_No": "",
        "Bill_Amount": "",
        "Incentive": "",
        "Ticket_Closed_On_Date": "",
        "Part_Name_Part_No": part_name or "",
        "Material_Description": material_desc,  # Auto-populated description
        "HSN_Code": "",
        "Quantity": "",
        "MRP_Product_Price": "",
        "CGST": "",
        "SGST": "",
        "Status": "OPEN",
        "Remarks": "",
        "Article_No": str(article_no).strip(),
        "Part_Sr_No": int(part_sr_no),
        "Money_Received": False,
        "Created_At": now,
        "Updated_At": now,
    }
    return pd.concat([df_jobs, pd.DataFrame([row])], ignore_index=True)


def get_service_jobs():
    migrate_db()
    df = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    return df


def update_service_job(row_id: str, updates: dict):
    """Update one billing row by Row_ID with validation."""
    migrate_db()
    df = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    if df.empty:
        return False, "No billing rows."

    rid = str(row_id).strip()
    mask = df["Row_ID"].astype(str).str.strip() == rid
    if not mask.any():
        return False, "Row not found."

    idx = df[mask].index[0]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Validation
    validation_errors = []

    # Check required fields when completing TCR
    tcr_no = updates.get("TCR_No", df.at[idx, "TCR_No"])
    if str(tcr_no).strip() and not str(tcr_no).strip():
        validation_errors.append("TCR Number cannot be empty when provided")

    # Check bill amount is numeric if provided
    bill_amt = updates.get("Bill_Amount", df.at[idx, "Bill_Amount"])
    if str(bill_amt).strip():
        try:
            float(str(bill_amt).strip())
        except ValueError:
            validation_errors.append("Bill Amount must be a valid number")

    # Check incentive is numeric if provided
    incentive = updates.get("Incentive", df.at[idx, "Incentive"])
    if str(incentive).strip():
        try:
            float(str(incentive).strip())
        except ValueError:
            validation_errors.append("Incentive Amount must be a valid number")

    if validation_errors:
        return False, "Validation errors: " + "; ".join(validation_errors)

    # Apply updates
    for k, v in (updates or {}).items():
        if k in SERVICE_JOB_COLUMNS and k != "Row_ID":
            df.at[idx, k] = v

    df.at[idx, "Updated_At"] = now

    # Enhanced status management
    tcr = df.at[idx, "TCR_No"]
    money = df.at[idx, "Money_Received"]
    bill_amt = df.at[idx, "Bill_Amount"]

    has_tcr = bool(str(tcr).strip()) if tcr is not None and str(tcr).strip().lower() != "nan" else False
    money_ok = money is True or str(money).strip().lower() in ("true", "1", "yes", "y", "received", "paid")
    has_bill = bool(str(bill_amt).strip()) if bill_amt is not None and str(bill_amt).strip().lower() != "nan" else False

    if has_tcr and money_ok:
        df.at[idx, "Status"] = "CLOSED"
        df.at[idx, "Ticket_Closed_On_Date"] = now.split()[0]  # Set close date
    elif has_tcr:
        df.at[idx, "Status"] = "COMPLETED"  # TCR submitted but not paid
    elif has_bill:
        df.at[idx, "Status"] = "IN_PROGRESS"  # Work done, bill raised
    else:
        df.at[idx, "Status"] = "OPEN"  # Just issued

    try:
        df_master = pd.read_excel(DB_FILE, sheet_name="Master", dtype={"Article_No": str})
    except Exception:
        df_master = pd.DataFrame(columns=MASTER_COLUMNS)
    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)
    df_eng = _read_sheet_df(ENGINEER_SHEET, ENGINEER_COLUMNS)

    save_all_workbook(df_master, df_trans, df_eng, df)
    return True, "Saved successfully."


def initialize_db():
    """Creates the file with all required sheets if it doesn't exist."""
    if not os.path.exists(DB_FILE):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        seeded = []
        for n in ENGINEERS:
            seeded.append({
                "Engineer_Name": n,
                "BP_ID": "",
                "PPRR_ID": "",
                "Active": True,
                "Created_At": now,
            })
        with pd.ExcelWriter(DB_FILE, engine='openpyxl') as writer:
            pd.DataFrame(columns=MASTER_COLUMNS).to_excel(writer, sheet_name='Master', index=False)
            pd.DataFrame(columns=TRANSACTION_COLUMNS).to_excel(writer, sheet_name='Transactions', index=False)
            pd.DataFrame(seeded, columns=ENGINEER_COLUMNS).to_excel(writer, sheet_name=ENGINEER_SHEET, index=False)
            pd.DataFrame(columns=SERVICE_JOB_COLUMNS).to_excel(writer, sheet_name=SERVICE_JOB_SHEET, index=False)
        print("✅ resQ Enterprise Database Initialized.")

def migrate_db():
    """
    Ensures existing Excel file has the latest schema/columns.
    Safe to call every startup.
    """
    if not os.path.exists(DB_FILE):
        initialize_db()
        return

    try:
        df_master = pd.read_excel(DB_FILE, sheet_name="Master", dtype={"Article_No": str})
    except Exception:
        df_master = pd.DataFrame(columns=MASTER_COLUMNS)

    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)

    # Engineers sheet (create if missing)
    try:
        df_eng = pd.read_excel(DB_FILE, sheet_name=ENGINEER_SHEET, dtype=str)
    except Exception:
        seeded = []
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for n in ENGINEERS:
            seeded.append({"Engineer_Name": n, "BP_ID": "", "PPRR_ID": "", "Active": True, "Created_At": now})
        df_eng = pd.DataFrame(seeded, columns=ENGINEER_COLUMNS)

    # Normalize columns and upgrade schema
    if not df_master.empty:
        for col in MASTER_COLUMNS:
            if col not in df_master.columns:
                df_master[col] = None
        df_master = df_master[MASTER_COLUMNS]
    else:
        df_master = pd.DataFrame(columns=MASTER_COLUMNS)

    df_trans = _ensure_schema(df_trans)

    # Normalize engineers schema
    if df_eng is None or df_eng.empty:
        df_eng = pd.DataFrame(columns=ENGINEER_COLUMNS)
    else:
        for col in ENGINEER_COLUMNS:
            if col not in df_eng.columns:
                df_eng[col] = None
        df_eng = df_eng[ENGINEER_COLUMNS]
        df_eng["Engineer_Name"] = df_eng["Engineer_Name"].astype(str).str.strip()
        # Ensure unique by name
        df_eng = df_eng.dropna(subset=["Engineer_Name"])
        df_eng = df_eng[df_eng["Engineer_Name"].astype(str).str.strip() != ""]
        df_eng = df_eng.drop_duplicates(subset=["Engineer_Name"], keep="last")

    try:
        df_jobs = pd.read_excel(DB_FILE, sheet_name=SERVICE_JOB_SHEET)
    except Exception:
        df_jobs = pd.DataFrame(columns=SERVICE_JOB_COLUMNS)

    df_jobs = _ensure_job_schema(df_jobs)

    with pd.ExcelWriter(DB_FILE, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master", index=False)
        df_trans.to_excel(writer, sheet_name="Transactions", index=False)
        df_eng.to_excel(writer, sheet_name=ENGINEER_SHEET, index=False)
        df_jobs.to_excel(writer, sheet_name=SERVICE_JOB_SHEET, index=False)

    apply_service_billing_green_rows()


def reset_database_keep_engineers() -> bool:
    """Reset all data sheets and preserve the Engineers sheet contents."""
    try:
        df_eng = pd.read_excel(DB_FILE, sheet_name=ENGINEER_SHEET, dtype=str)
    except Exception:
        df_eng = None

    if df_eng is not None and not df_eng.empty:
        for col in ENGINEER_COLUMNS:
            if col not in df_eng.columns:
                df_eng[col] = None
        df_eng = df_eng[ENGINEER_COLUMNS].copy()
        df_eng["Engineer_Name"] = df_eng["Engineer_Name"].astype(str).str.strip()
        df_eng = df_eng[df_eng["Engineer_Name"] != ""]
        df_eng = df_eng.drop_duplicates(subset=["Engineer_Name"], keep="last")
    else:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        seeded = [
            {
                "Engineer_Name": n,
                "BP_ID": "",
                "PPRR_ID": "",
                "Active": True,
                "Created_At": now,
            }
            for n in ENGINEERS
        ]
        df_eng = pd.DataFrame(seeded, columns=ENGINEER_COLUMNS)

    df_master = pd.DataFrame(columns=MASTER_COLUMNS)
    df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)
    df_jobs = pd.DataFrame(columns=SERVICE_JOB_COLUMNS)

    save_all_workbook(df_master, df_trans, df_eng, df_jobs)
    return True


def get_engineers(active_only: bool = True):
    """Return engineers as list of dicts: name + ids."""
    migrate_db()
    try:
        df = pd.read_excel(DB_FILE, sheet_name=ENGINEER_SHEET, dtype=str)
    except Exception:
        return []
    if df is None or df.empty:
        return []
    for col in ENGINEER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[ENGINEER_COLUMNS].copy()
    df["Engineer_Name"] = df["Engineer_Name"].astype(str).str.strip()
    df = df[df["Engineer_Name"] != ""]
    if active_only and "Active" in df.columns:
        # Accept True/False or "True"/"False"
        df["Active"] = df["Active"].astype(str).str.lower().isin(["true", "1", "yes", "y"])
        df = df[df["Active"] == True]
    out = []
    for _, r in df.iterrows():
        out.append({
            "Engineer_Name": str(r.get("Engineer_Name", "")).strip(),
            "BP_ID": "" if r.get("BP_ID") is None else str(r.get("BP_ID")).strip(),
            "PPRR_ID": "" if r.get("PPRR_ID") is None else str(r.get("PPRR_ID")).strip(),
        })
    return out

def get_engineer_names(active_only: bool = True):
    return [e["Engineer_Name"] for e in get_engineers(active_only=active_only)]

def add_engineer(name: str, bp_id: str = "", pprr_id: str = ""):
    """Add or update an engineer in Engineers sheet."""
    migrate_db()
    name = "" if name is None else str(name).strip()
    if not name:
        return False, "Engineer name is required."

    bp_id = "" if bp_id is None else str(bp_id).strip()
    pprr_id = "" if pprr_id is None else str(pprr_id).strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        df = pd.read_excel(DB_FILE, sheet_name=ENGINEER_SHEET, dtype=str)
    except Exception:
        df = pd.DataFrame(columns=ENGINEER_COLUMNS)

    for col in ENGINEER_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df["Engineer_Name"] = df["Engineer_Name"].astype(str).str.strip()
    mask = df["Engineer_Name"].str.lower() == name.lower()
    if mask.any():
        idx = df[mask].index[0]
        df.at[idx, "Engineer_Name"] = name
        df.at[idx, "BP_ID"] = bp_id
        df.at[idx, "PPRR_ID"] = pprr_id
        df.at[idx, "Active"] = True
        msg = f"Updated engineer: {name}"
    else:
        df = pd.concat([df, pd.DataFrame([{
            "Engineer_Name": name,
            "BP_ID": bp_id,
            "PPRR_ID": pprr_id,
            "Active": True,
            "Created_At": now,
        }])], ignore_index=True)
        msg = f"Added engineer: {name}"

    # Keep unique
    df = df.drop_duplicates(subset=["Engineer_Name"], keep="last")

    # Write back all sheets safely
    try:
        df_master = pd.read_excel(DB_FILE, sheet_name="Master", dtype={"Article_No": str})
    except Exception:
        df_master = pd.DataFrame(columns=MASTER_COLUMNS)
    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)
    df_trans = _ensure_schema(df_trans)

    df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)

    with pd.ExcelWriter(DB_FILE, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master", index=False)
        df_trans.to_excel(writer, sheet_name="Transactions", index=False)
        df[ENGINEER_COLUMNS].to_excel(writer, sheet_name=ENGINEER_SHEET, index=False)
        df_jobs.to_excel(writer, sheet_name=SERVICE_JOB_SHEET, index=False)

    apply_service_billing_green_rows()

    return True, msg

def remove_engineer(name: str):
    """Soft-remove engineer by setting Active=False."""
    migrate_db()
    name = "" if name is None else str(name).strip()
    if not name:
        return False, "Engineer name is required."

    try:
        df = pd.read_excel(DB_FILE, sheet_name=ENGINEER_SHEET, dtype=str)
    except Exception:
        return False, "Engineers sheet not found."

    for col in ENGINEER_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df["Engineer_Name"] = df["Engineer_Name"].astype(str).str.strip()
    mask = df["Engineer_Name"].str.lower() == name.lower()
    if not mask.any():
        return False, f"Engineer not found: {name}"

    idx = df[mask].index[0]
    df.at[idx, "Active"] = False

    # Write back all sheets safely
    try:
        df_master = pd.read_excel(DB_FILE, sheet_name="Master", dtype={"Article_No": str})
    except Exception:
        df_master = pd.DataFrame(columns=MASTER_COLUMNS)
    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)
    df_trans = _ensure_schema(df_trans)

    df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)

    with pd.ExcelWriter(DB_FILE, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master", index=False)
        df_trans.to_excel(writer, sheet_name="Transactions", index=False)
        df[ENGINEER_COLUMNS].to_excel(writer, sheet_name=ENGINEER_SHEET, index=False)
        df_jobs.to_excel(writer, sheet_name=SERVICE_JOB_SHEET, index=False)

    apply_service_billing_green_rows()

    return True, f"Removed engineer: {name}"

def _ensure_schema(df_trans: pd.DataFrame) -> pd.DataFrame:
    """
    Backward-compatible schema upgrade:
    - Old files may not have `Sr_No`. Default existing rows to Sr_No=1.
    - Ensure all expected columns exist.
    """
    if df_trans is None or df_trans.empty:
        return pd.DataFrame(columns=TRANSACTION_COLUMNS)

    df = df_trans.copy()
    if "Sr_No" not in df.columns:
        df.insert(1, "Sr_No", 1)

    for col in TRANSACTION_COLUMNS:
        if col not in df.columns:
            df[col] = None

    # If we are upgrading from older schema that had only `Date`,
    # populate missing In_Date / Out_Date based on current Status.
    if "Date" in df.columns:
        if "In_Date" in df.columns:
            in_missing = df["In_Date"].isna() | (df["In_Date"].astype(str).str.strip() == "") | (df["In_Date"].astype(str) == "nan")
            df.loc[in_missing & (df["Status"] == "IN"), "In_Date"] = df.loc[in_missing & (df["Status"] == "IN"), "Date"]
        if "Out_Date" in df.columns:
            out_missing = df["Out_Date"].isna() | (df["Out_Date"].astype(str).str.strip() == "") | (df["Out_Date"].astype(str) == "nan")
            df.loc[out_missing & (df["Status"] == "OUT"), "Out_Date"] = df.loc[out_missing & (df["Status"] == "OUT"), "Date"]

    return df[TRANSACTION_COLUMNS]

def _normalize_ids(df_master: pd.DataFrame, df_trans: pd.DataFrame):
    df_master["Article_No"] = df_master["Article_No"].astype(str).str.strip()
    df_trans["Article_No"] = df_trans["Article_No"].astype(str).str.strip()
    df_trans["Sr_No"] = pd.to_numeric(df_trans["Sr_No"], errors="coerce").fillna(1).astype(int)
    return df_master, df_trans

def get_available_sr_nos(article_no: str):
    """Returns Sr_No list that are currently IN for an Article_No."""
    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        return []

    df_trans = _ensure_schema(df_trans)
    df_trans["Article_No"] = df_trans["Article_No"].astype(str).str.strip()
    df_trans["Sr_No"] = pd.to_numeric(df_trans["Sr_No"], errors="coerce").fillna(1).astype(int)

    search_id = str(article_no).strip()
    active_mask = (df_trans["Article_No"] == search_id) & (df_trans["Status"] == "IN")
    srs = sorted(df_trans.loc[active_mask, "Sr_No"].astype(int).unique().tolist())
    return srs

def get_scan_details(article_no: str):
    """
    Returns details to show after scanning:
    - Article_No, Part_Name, SP
    - In-units list: [{Sr_No, Tax_Invoice_No}]
    """
    search_id = str(article_no).strip()
    try:
        df_master = pd.read_excel(DB_FILE, sheet_name="Master", dtype={"Article_No": str})
    except Exception:
        return {
            "Article_No": search_id,
            "Part_Name": "",
            "SP": None,
            "in_units": [],
        }

    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)

    df_trans = _ensure_schema(df_trans)
    df_master, df_trans = _normalize_ids(df_master, df_trans)

    part_name = ""
    sp = None
    if search_id in df_master["Article_No"].values:
        idx = df_master.index[df_master["Article_No"] == search_id][0]
        part_name = df_master.at[idx, "Part_Name"]
        sp = df_master.at[idx, "SP"]

    active_mask = (df_trans["Article_No"] == search_id) & (df_trans["Status"] == "IN")
    active = df_trans.loc[active_mask, ["Sr_No", "Tax_Invoice_No"]].copy()
    if not active.empty:
        active["Sr_No"] = pd.to_numeric(active["Sr_No"], errors="coerce").fillna(1).astype(int)
        active = active.sort_values("Sr_No")

    in_units = []
    for _, r in active.iterrows():
        inv = r.get("Tax_Invoice_No", "")
        inv = "" if inv is None else str(inv).strip()
        in_units.append({"Sr_No": int(r["Sr_No"]), "Tax_Invoice_No": inv})

    return {
        "Article_No": search_id,
        "Part_Name": part_name,
        "SP": sp,
        "in_units": in_units,
    }

def get_next_inward_sr_list(article_no: str, qty: int):
    """Returns the Sr_No list that will be allocated by auto-inward logic."""
    search_id = str(article_no).strip()
    try:
        qty_int = int(qty)
    except Exception:
        qty_int = 1
    if qty_int < 1:
        qty_int = 1

    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name="Transactions", dtype={"Article_No": str})
    except Exception:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)

    df_trans = _ensure_schema(df_trans)
    df_trans["Article_No"] = df_trans["Article_No"].astype(str).str.strip()
    df_trans["Sr_No"] = pd.to_numeric(df_trans["Sr_No"], errors="coerce").fillna(0).astype(int)

    existing_for_article = df_trans[df_trans["Article_No"] == search_id]
    max_sr = int(existing_for_article["Sr_No"].max()) if not existing_for_article.empty else 0
    return list(range(max_sr + 1, max_sr + qty_int + 1))

def register_new_article(art_no, name, cp, sp, cat="OG"):
    # Load all existing data first to prevent losing sheets
    df_master = pd.read_excel(DB_FILE, sheet_name='Master', dtype={'Article_No': str})
    try:
        df_trans = pd.read_excel(DB_FILE, sheet_name='Transactions', dtype={'Article_No': str})
    except:
        df_trans = pd.DataFrame(columns=TRANSACTION_COLUMNS)

    df_trans = _ensure_schema(df_trans)

    art_no = str(art_no).strip()
    if art_no in df_master['Article_No'].astype(str).str.strip().values:
        return False, f"❌ Article '{art_no}' already exists!"
    
    new_entry = pd.DataFrame([{
        "Article_No": art_no, "Part_Name": name, "CP": cp, "SP": sp, "Category": cat, "Stock_Level": 0
    }])
    
    df_master = pd.concat([df_master, new_entry], ignore_index=True)

    df_eng = _read_sheet_df(ENGINEER_SHEET, ENGINEER_COLUMNS)
    df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    save_all_workbook(df_master, df_trans, df_eng, df_jobs)

    return True, f"✅ Article {art_no} registered."

def process_movement(art_no, movement_type, purchase_type="Company", engineer=None, sr_no=None, qty: int = 1, tax_invoice_no: str = ""):
    try:
        df_master = pd.read_excel(DB_FILE, sheet_name='Master', dtype={'Article_No': str})
        df_trans = pd.read_excel(DB_FILE, sheet_name='Transactions', dtype={'Article_No': str})
    except:
        return False, "❌ Error reading database."

    df_trans = _ensure_schema(df_trans)

    search_id = str(art_no).strip()
    df_master, df_trans = _normalize_ids(df_master, df_trans)

    if search_id not in df_master['Article_No'].values:
        return False, f"❌ Article '{search_id}' not found in Master!"

    master_idx = df_master.index[df_master['Article_No'] == search_id][0]
    product_name = df_master.at[master_idx, 'Part_Name']

    append_out_job = False
    out_ts_for_job = None

    if movement_type.upper() == "IN":
        try:
            qty_int = int(qty)
        except Exception:
            qty_int = 1
        if qty_int < 1:
            qty_int = 1

        # Auto-assign Sr_No sequence if not provided.
        if sr_no is not None:
            sr_list = [int(sr_no)]
        else:
            existing_for_article = df_trans[df_trans["Article_No"] == search_id]
            max_sr = int(existing_for_article["Sr_No"].max()) if not existing_for_article.empty else 0
            sr_list = list(range(max_sr + 1, max_sr + qty_int + 1))

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        added_or_reset = 0
        tax_invoice_no = "" if tax_invoice_no is None else str(tax_invoice_no).strip()

        for sr in sr_list:
            unit_mask = (df_trans["Article_No"] == search_id) & (df_trans["Sr_No"] == int(sr))

            if unit_mask.any():
                # If this unit is already IN, don't duplicate it.
                current_status = df_trans.loc[unit_mask, "Status"].values[0]
                if current_status == "IN":
                    continue
                # If OUT, allow re-entry/reset to IN
                target_idx = df_trans[unit_mask].index[0]
                df_trans.at[target_idx, "Status"] = "IN"
                df_trans.at[target_idx, "Purchase_Type"] = purchase_type
                df_trans.at[target_idx, "Tax_Invoice_No"] = tax_invoice_no
                df_trans.at[target_idx, "Engineer"] = "Store"
                df_trans.at[target_idx, "Charges"] = 0
                df_trans.at[target_idx, "In_Date"] = now
                df_trans.at[target_idx, "Out_Date"] = None
                df_trans.at[target_idx, "Date"] = now
                added_or_reset += 1
            else:
                new_log = pd.DataFrame([{
                    "Article_No": search_id,
                    "Sr_No": int(sr),
                    "Purchase_Type": purchase_type,
                    "Tax_Invoice_No": tax_invoice_no,
                    "Status": "IN",
                    "Engineer": "Store",
                    "Charges": 0,
                    "In_Date": now,
                    "Out_Date": None,
                    "Date": now
                }])
                df_trans = pd.concat([df_trans, new_log], ignore_index=True)
                added_or_reset += 1

        # Update Master stock = count of IN units for this article
        in_count = int(((df_trans["Article_No"] == search_id) & (df_trans["Status"] == "IN")).sum())
        df_master.at[master_idx, "Stock_Level"] = in_count

        if len(sr_list) == 1:
            msg = f"✅ INWARD: {product_name} ({search_id}) Sr#{sr_list[0]} added to store."
        else:
            msg = f"✅ INWARD: {product_name} ({search_id}) added {added_or_reset}/{len(sr_list)} unit(s) to store."

    elif movement_type.upper() == "OUT":
        if engineer is None or str(engineer).strip() == "":
            return False, "❌ Select Engineer first."

        # If Sr_No not given, auto-pick if only one unit is IN, otherwise block.
        active_mask = (df_trans["Article_No"] == search_id) & (df_trans["Status"] == "IN")
        active_srs = sorted(df_trans.loc[active_mask, "Sr_No"].astype(int).unique().tolist())

        if not active_srs:
            return False, f"❌ ERROR: {search_id} is not in store (all units OUT)."

        if sr_no is None:
            if len(active_srs) == 1:
                sr_no = active_srs[0]
            else:
                return False, f"⚠️ Multiple units IN for {search_id}. Select Sr No: {active_srs}"

        unit_mask = (df_trans["Article_No"] == search_id) & (df_trans["Sr_No"] == int(sr_no)) & (df_trans["Status"] == "IN")
        if unit_mask.any():
            target_idx = df_trans[unit_mask].index[0]
            df_trans.at[target_idx, "Status"] = "OUT"
            df_trans.at[target_idx, "Engineer"] = engineer
            df_trans.at[target_idx, "Charges"] = df_master.at[master_idx, "SP"]
            out_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            df_trans.at[target_idx, "Out_Date"] = out_now
            df_trans.at[target_idx, "Date"] = out_now

            in_count = int(((df_trans["Article_No"] == search_id) & (df_trans["Status"] == "IN")).sum())
            df_master.at[master_idx, "Stock_Level"] = in_count
            msg = f"📤 ISSUED: {product_name} ({search_id}) Sr#{int(sr_no)} given to {engineer}."
            append_out_job = True
            out_ts_for_job = out_now
        else:
            return False, f"❌ ERROR: {search_id} Sr#{int(sr_no)} is not IN store (might be already OUT)."

    df_eng = _read_sheet_df(ENGINEER_SHEET, ENGINEER_COLUMNS)
    df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    if append_out_job:
        df_jobs = append_service_job_after_outward(
            df_jobs,
            article_no=search_id,
            part_sr_no=int(sr_no),
            engineer_name=str(engineer).strip(),
            part_name=str(product_name),
            outward_dt=str(out_ts_for_job),
        )

    save_all_workbook(df_master, df_trans, df_eng, df_jobs)
    return True, msg


# ===== TCR TRACKING SEARCH FUNCTIONS =====

def search_tcr_by_engineer(engineer_name: str):
    """
    Search TCR records by engineer name.
    Returns both complete (TCR filled) and incomplete (no TCR) records for the engineer.
    """
    migrate_db()
    engineer_name = str(engineer_name).strip()
    
    try:
        df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    except Exception:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    
    if df_jobs.empty:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    
    # Case-insensitive search
    df_jobs["Engineer_Name"] = df_jobs["Engineer_Name"].astype(str).str.strip()
    mask = df_jobs["Engineer_Name"].str.lower() == engineer_name.lower()
    
    result = df_jobs[mask].copy()
    # Sort by creation date, newest first
    result = result.sort_values("Created_At", ascending=False, na_position="last")
    
    return result


def search_tcr_by_article(article_no: str):
    """
    Search TCR records by article number.
    Returns both complete and incomplete engineer parts list for the article.
    """
    migrate_db()
    article_no = str(article_no).strip()
    
    try:
        df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
    except Exception:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    
    if df_jobs.empty:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    
    # Search by article number
    df_jobs["Article_No"] = df_jobs["Article_No"].astype(str).str.strip()
    mask = df_jobs["Article_No"] == article_no
    
    result = df_jobs[mask].copy()
    # Sort by engineer name then by creation date
    result = result.sort_values(["Engineer_Name", "Created_At"], ascending=[True, False])
    
    return result


def search_tcr_by_artcode(art_code: str):
    """
    Search TCR records by article code (Art Code).
    Returns list of both complete and incomplete engineer parts.
    Art Code is the article number identifier used in the system.
    """
    # Art code is essentially the Article_No, so use the same logic
    return search_tcr_by_article(art_code)


def filter_tcr_records(df_jobs: pd.DataFrame, search_type: str = None, search_value: str = None, status_filter: str = "ALL"):
    """
    Filter TCR records with optional search and status filtering.
    
    search_type: 'engineer', 'article', 'artcode', or None
    search_value: the value to search for
    status_filter: 'ALL', 'OPEN', 'IN_PROGRESS', 'COMPLETED', 'CLOSED'
    """
    if df_jobs is None or df_jobs.empty:
        return pd.DataFrame(columns=SERVICE_JOB_COLUMNS)
    
    result = df_jobs.copy()
    
    # Apply search filter
    if search_type and search_value:
        search_value = str(search_value).strip()
        
        if search_type.lower() == "engineer":
            result["Engineer_Name"] = result["Engineer_Name"].astype(str).str.strip()
            result = result[result["Engineer_Name"].str.lower().str.contains(search_value.lower(), na=False)]
        
        elif search_type.lower() == "article":
            result["Article_No"] = result["Article_No"].astype(str).str.strip()
            result = result[result["Article_No"].str.contains(search_value, na=False)]
        
        elif search_type.lower() == "artcode":
            result["Article_No"] = result["Article_No"].astype(str).str.strip()
            result = result[result["Article_No"].str.contains(search_value, na=False)]
    
    # Apply status filter
    if status_filter and status_filter.upper() != "ALL":
        result["Status"] = result["Status"].astype(str).str.strip().str.upper()
        result = result[result["Status"] == status_filter.upper()]
    
    # Sort by creation date, newest first
    result = result.sort_values("Created_At", ascending=False, na_position="last")
    
    return result


def get_tcr_completion_stats(df_jobs: pd.DataFrame = None):
    """
    Get statistics about TCR completion status.
    Returns a dictionary with counts of different statuses.
    """
    if df_jobs is None:
        try:
            df_jobs = _read_sheet_df(SERVICE_JOB_SHEET, SERVICE_JOB_COLUMNS)
        except Exception:
            return {"total": 0, "OPEN": 0, "IN_PROGRESS": 0, "COMPLETED": 0, "CLOSED": 0}
    
    if df_jobs.empty:
        return {"total": 0, "OPEN": 0, "IN_PROGRESS": 0, "COMPLETED": 0, "CLOSED": 0}
    
    df_jobs["Status"] = df_jobs["Status"].astype(str).str.strip().str.upper()
    
    stats = {
        "total": len(df_jobs),
        "OPEN": int((df_jobs["Status"] == "OPEN").sum()),
        "IN_PROGRESS": int((df_jobs["Status"] == "IN_PROGRESS").sum()),
        "COMPLETED": int((df_jobs["Status"] == "COMPLETED").sum()),
        "CLOSED": int((df_jobs["Status"] == "CLOSED").sum()),
    }
    
    return stats


def reformat_excel_file():
    """
    Reformat the Excel file with auto-sized columns and professional formatting.
    Can be called anytime to improve readability.
    """
    migrate_db()
    apply_excel_formatting()
    return True, "✅ Excel file reformatted with auto-sized columns."